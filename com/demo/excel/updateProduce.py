#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time  : 2020/8/24 19:44
# @Author: yuzq
# @File  : updateProduce
import gc
import io
import os

import openpyxl
from openpyxl.chart import reference, series, title, BarChart, Reference
from openpyxl.chart.series import Series
from openpyxl.comments import Comment
from openpyxl.drawing import Drawing
from openpyxl.styles import Font, Alignment
from openpyxl.utils import column_index_from_string

if __name__ == '__main__':
    book_dir = r'E:\lang\python\demo\com\books\Automate_the_Boring_Stuff_onlinematerials-master'
    file_path = os.path.join(book_dir, 'updateProduce.xlsx')

    # with open('updatedProduceSales.xlsx', "rb") as f:
    #     in_mem_file = io.BytesIO(f.read())

    wb = openpyxl.load_workbook('updatedProduceSales1.xlsx')
    sheet = wb.active

    PRICE_UPDATES = {
        'Garlic': 3.07,
        'Celery': 1.19,
        'Lemon': 1.27
    }

    # 设置行高和列宽
    sheet.row_dimensions[1].height = 50
    sheet.column_dimensions['A'].width = 50

    # 合并单元格
    # sheet.merge_cells('A1:D3')
    # sheet['A1'] = 'merge message'

    # 拆分单元格
    # sheet.unmerge_cells('A1:D3')

    # 冻结窗格
    sheet.freeze_panes = 'B2'

    # 创建图表
    col_num = column_index_from_string('F')
    print(col_num)

    data = Reference(sheet, min_row=2, max_row=13, min_col=col_num, max_col=col_num)
    cats = Reference(sheet, min_row=2, max_row=13, min_col=col_num)

    bar_chart = BarChart()
    bar_chart.type='col'
    bar_chart.style=10
    bar_chart.title = 'Bar Chart'
    bar_chart.x_axis.title='X轴'
    bar_chart.y_axis.title = 'y轴'
    bar_chart.add_data(data)
    bar_chart.set_categories(cats)
    # drawing = Drawing()
    # drawing.top = 50
    # drawing.left = 100
    # drawing.width = 300
    # drawing.height = 200
    sheet.add_chart(bar_chart,'H5')


    # 字体样式
    italic24_font = Font(size=24, italic=True, name='Times New Roman', bold=True)

    comment = Comment('批注',author='yu')

    for row_num in range(2, sheet.max_row):
        produce_name = sheet.cell(row=row_num, column=1).value

        price = sheet.cell(row=row_num, column=2).value

        sold = sheet.cell(row=row_num, column=3).value

        # 添加批注
        # sheet.cell(row=row_num, column=1).comment = comment

        if (produce_name in PRICE_UPDATES):
            sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]

        # if price > 5.0:
        #     sheet.cell(row=row_num, column=1).font = italic24_font

        # 设置公式
        sheet.cell(row=row_num, column=6).value = '=SUM(B' + str(row_num) + ':C' + str(row_num) + ')'

        # 设置对齐方式
        sheet.cell(row=row_num,column=6).alignment = Alignment(horizontal='center',vertical='center')


    wb.save('updatedProduceSales2.xlsx')

    del wb, sheet  # wb为打开的工作表
    gc.collect()
