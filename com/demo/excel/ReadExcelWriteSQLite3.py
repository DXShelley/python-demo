#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time  : 2020/8/20 15:42
# @Author: yuzq
# @File  : ReadExcelWriteSQLite3
import datetime
import logging
import os
import sqlite3

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

logging.basicConfig(level=logging.DEBUG)


def read_excel(file_path,col_names=[]):
    logging.debug('开始读取excel文件。。。')
    workbook = openpyxl.load_workbook(file_path)  # type:openpyxl.workbook.workbook.Workbook
    # logging.debug(type(workbook))
    sheet = workbook.active  # type:openpyxl.worksheet.worksheet.Worksheet
    # logging.debug(type(sheet))
    # logging.debug(sheet.title)

    # logging.debug(type(sheet['A2']))
    # logging.debug(sheet['A2'].value)
    # logging.debug(type(sheet['A2'].value))
    # logging.debug(sheet.cell(2, 1).value)
    # logging.debug(sheet.max_row)
    # logging.debug(sheet.max_column)
    #
    # logging.debug(get_column_letter(1))
    # logging.debug(column_index_from_string('A'))

    # min_row = sheet.min_row
    # min_col = sheet.min_column
    #
    # first_row = []
    # for colObject in sheet.columns[min_col:]:
    #     logging.debug(colObject)
    #     first_row.append( [row for row in colObject if(row.row == min_row)])
    #
    # col_names = list(set(col_names).intersection(first_row))

    rows = []
    for row_object in sheet.rows:
        row = []
        if row_object[0].row == 1:
            continue
        for col_object in row_object:
            if isinstance(col_object.value, datetime.datetime):
                value = col_object.value.strftime("%Y-%m-%d %H:%M:%S")
                row.append(value)
            else:
                row.append(col_object.value)
        rows.append(tuple(row))
    # logging.debug(list(rows))
    logging.debug('读取excel文件成功。')
    return rows

def write_sqlite3(rows,col_names=None):
    logging.debug('开始导入sqlite3。。。')
    con = sqlite3.connect(r'E:\lang\python\demo\com\demo\db\sqlite\example.db')
    cursor = con.cursor()
    insert_tables = '''INSERT INTO stocks (date, trans, symbol, qty, price) VALUES (?,?,?,?,?);'''
    cursor.executemany(insert_tables, rows)
    con.commit()

    cursor.close()
    con.close()
    logging.debug('导入成功。')


if __name__ == '__main__':
    book_dir = r'E:\lang\python\demo\com\books\Automate_the_Boring_Stuff_onlinematerials-master'
    file_path = os.path.join(book_dir, 'example1.xlsx')
    rows = read_excel(file_path)
    write_sqlite3(rows)

