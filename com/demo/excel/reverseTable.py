#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time  : 2020/8/29 16:35
# @Author: yuzq
# @File  : reverseTable

import openpyxl
from openpyxl.workbook import workbook

if __name__ == '__main__':
    wb = openpyxl.load_workbook('updatedProduceSales2.xlsx')
    sheet = wb.active

    cells = {}
    for x in range(1, sheet.max_row):
        cells.setdefault(x, {})
        for y in range(1, sheet.max_column):
            cells[x].setdefault(y, sheet.cell(row=x, column=y).value)

    wb2 = workbook.Workbook()
    sheet2 = wb2.create_sheet('Reversed', 0)
    for x in range(1, len(cells.keys())):
        for y in range(1, len(cells.get(x))):
            sheet2.cell(row=y, column=x).value = cells[x][y]

    # 行转列 公式计算错误 todo
    wb2.save('updatedProduceSales3.xlsx')
    wb.close()
    wb2.close()
